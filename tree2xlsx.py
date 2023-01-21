#################################################
#                                               #
# Tree2XLSX - File tree structure to XLSX table #
#                                               #
#################################################

import glob
import sys

from openpyxl import Workbook,load_workbook

def _help():
	print('''
Tree2XLSX - File&Folder tree structure to XLSX table.

Usage: python3 tree2xlsx.py [working directory] [table file]
Example: python3 tree2xlsx.py '/home' ./files_folders.xlsx

version 1.0
developed by memeister
updated 2023-01-21
''')

def gettree():
	# Recursively retrieve files and folders throughout specified working directory
	ff = glob.glob(f"{sys.argv[1]}/**")
	return ff

def addtoxlsx(ff):
	for x in ff:
		exist = False
		for y in range(ws.max_row):
			if(ws[f'A{y+1}'].internal_value == x):
				exist = True
				break
		if(exist == False):
			ws.append([x,x.split('/')[-1]]) # Append filename,filepath
		else:
			continue

if(__name__ == "__main__"):
	if('--help' in sys.argv or len(sys.argv)-1 < 2):
		_help()
		sys.exit(0)
	else:
		try:
			wb = load_workbook(filename = sys.argv[2])
		except FileNotFoundError:
			print(f"File not found '{sys.argv[2]}' .. creating new file!")
			wb = Workbook()
		ws = wb.active
		tree = gettree()
		print(f"Found {len(tree)} files and folders in '{sys.argv[1]}'")
		addtoxlsx(tree)
		wb.save(sys.argv[2])
		print(f"Written to file '{sys.argv[2]}'")
		sys.exit(0)
	sys.exit(1)
